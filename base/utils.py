import os
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill

def calculate_component(amount_trk, amount_eis, amount_vts, amount):
    spent_eis = 0
    spent_trk = 0
    spent_vts = 0
    result = {}
    while amount_trk > 0 and amount > 0:
        amount -= 1
        amount_trk -= 1
        spent_trk += 1
    result["amount_trk"] = amount_trk
    result["spent_trk"] = spent_trk

    while amount_eis > 0 and amount > 0:
        amount -= 1
        amount_eis -= 1
        spent_eis += 1
    result["amount_eis"] = amount_eis
    result["spent_eis"] = spent_eis

    while amount_vts > 0 and amount > 0:
        amount -= 1
        amount_vts -= 1
        spent_vts += 1
    result["amount_vts"] = amount_vts
    result["spent_vts"] = spent_vts
    result["amount"] = amount
    return result


def create_statement(block, cleaned_data, date):
    path_template = os.path.join(
        os.getcwd(), "base/Statement/Defect.xlsx"
    )
    wb = load_workbook(filename=path_template)
    dest_filename = (
        f"{block.number_block}_{block.serial_number}_{block.name_block}_{block.region}.xlsx"
    )
    sheet = wb["Лист1"]
    sheet["B5"].value = block.number_block
    sheet["D5"].value = date
    sheet["A17"].value = str(block.name_block)
    sheet["I20"].value = block.serial_number
    sheet["C20"].value = str(block.region)
    sheet["B26"].value = cleaned_data['defect_1']
    sheet["B28"].value = cleaned_data['defect_2']
    sheet["B30"].value = cleaned_data['defect_3']
    sheet["C47"].value = cleaned_data['result']
    path_save = os.path.join(
        os.getcwd(),
        "base/Statement/Defects/{}".format(
            dest_filename.replace("/", "")
        ),
    )
    wb.save(path_save)


def create_repair_maker(qs):
    path_template = os.path.join(
        os.getcwd(), "base/Maker/pattern.xlsx"
    )
    wb = load_workbook(filename=path_template)
    sheet = wb["Лист1"]
    cnt = 2
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    for record in qs:
        sheet["A" + str(cnt)].value = str(record.block.number_block)
        sheet["B" + str(cnt)].value = str(record.block.name_block)
        sheet["C" + str(cnt)].value = str(record.block.serial_number)
        sheet["D" + str(cnt)].value = str(record.block.region)
        sheet["E" + str(cnt)].value = str(record.maker)
        sheet["F" + str(cnt)].value = record.block.date_add
        sheet["G" + str(cnt)].value = record.date_shipment_maker
        sheet["H" + str(cnt)].value = record.date_add_maker
        sheet["I" + str(cnt)].value = record.maker_status
        cnt += 1
    for i in range(1, len(qs)+2):
        sheet.cell(row=i, column=1).border = thin_border
        sheet.cell(row=i, column=2).border = thin_border
        sheet.cell(row=i, column=3).border = thin_border
        sheet.cell(row=i, column=4).border = thin_border
        sheet.cell(row=i, column=5).border = thin_border
        sheet.cell(row=i, column=6).border = thin_border
        sheet.cell(row=i, column=7).border = thin_border
        sheet.cell(row=i, column=8).border = thin_border
        sheet.cell(row=i, column=9).border = thin_border
    path_save = os.path.join(
        os.getcwd(),
        "base/Maker/repair_maker.xlsx"
    )
    wb.save(path_save)


def create_report_excel(order, date_after, date_before, company, unit_order, purpose_order):
    unique_invoice_number = list({v['invoice_number']:v for v in order.values('invoice_number', 'invoice_amount', 'provider', 'unit_order', 'purpose_order')}.values()) # получаем уникальные номера счетов и сумму счетов
    invoice_order = {}
    unique_unit_order = []
    all_cnt_component = 0
    all_summ_invoice = 0
    for i in unique_invoice_number:
        cnt_component = order.filter(invoice_number=i['invoice_number']).values('component').count() # считаем количество компонентов в счете
        invoice_order[i['invoice_number']] = [cnt_component, i['invoice_amount']] # словарь с указанием номера счета, количества компонентов в счете и сумма счета
        #unit_order[i['unit_order']] = [i['invoice_number'], cnt_component, i['invoice_amount'], i['provider'], i['unit_order'], i['purpose_order']]
        unique_unit_order.append([i['invoice_number'], cnt_component, i['invoice_amount'], i['provider'], i['unit_order'], i['purpose_order']])
        all_cnt_component += cnt_component
        all_summ_invoice += i['invoice_amount']

    agr = []
    for i in unique_unit_order:
        agr.append(i[2])
    if agr:
        max_invoice = max(agr)
        min_invoice = min(agr)
        avg_invoice = sum(agr) / len(agr)

    path_template = os.path.join(
        os.getcwd(), "base/Order_report/report_order_template.xlsx"
    )
    wb = load_workbook(filename=path_template)
    sheet = wb["Лист1"]
    cnt = 5

    for record in order:
        sheet["B" + str(cnt)].value = str(record.invoice_number)
        sheet["C" + str(cnt)].value = str(record.component.type_component)
        sheet["D" + str(cnt)].value = str(record.component.marking)
        sheet["E" + str(cnt)].value = str(record.amount_commit)
        sheet["B" + str(cnt)].alignment = Alignment(horizontal='center')
        sheet["E" + str(cnt)].alignment = Alignment(horizontal='center')
        cnt += 1
    
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    head = Font(name='Times New Roman', size=16)
    fill = PatternFill(
        fill_type='solid',
        fgColor="DDDDDD"
    )
    cnt = 5
    for i in unit_order:
        sheet["G" + str(cnt)].font = head
        sheet["G" + str(cnt)].fill = fill
        sheet["G" + str(cnt)].value = str(i)
        sheet["G" + str(cnt)].alignment = Alignment(horizontal='center')
        sheet.merge_cells(f'G{cnt}:K{cnt}')
        cnt += 1
        sum_purpose = 0
        for j in unique_unit_order:
            if i == unit_order.get(pk=j[4]):
                if j[5]:
                    purpose = purpose_order.get(pk=j[5])
                else:
                    purpose = None
                sheet["G" + str(cnt)].value = str(j[0])
                sheet["H" + str(cnt)].value = int(j[1])
                sheet["I" + str(cnt)].value = str(j[3])
                sheet["J" + str(cnt)].value = float(j[2])
                sheet["K" + str(cnt)].value = str(purpose)
                sheet["G" + str(cnt)].alignment = Alignment(horizontal='center')
                sheet["H" + str(cnt)].alignment = Alignment(horizontal='center')
                sheet["I" + str(cnt)].alignment = Alignment(horizontal='center')
                sheet["J" + str(cnt)].alignment = Alignment(horizontal='center')
                sheet["K" + str(cnt)].alignment = Alignment(horizontal='center')
                sheet.cell(row=cnt, column=7).border = thin_border
                sheet.cell(row=cnt, column=8).border = thin_border
                sheet.cell(row=cnt, column=9).border = thin_border
                sheet.cell(row=cnt, column=10).border = thin_border
                sheet.cell(row=cnt, column=11).border = thin_border
                cnt += 1
                sum_purpose += j[2]
        sheet["I" + str(cnt)].value = 'Итого:'
        sheet["J" + str(cnt)].value = float(sum_purpose)
        sheet["I" + str(cnt)].alignment = Alignment(horizontal='center')
        sheet["J" + str(cnt)].alignment = Alignment(horizontal='center')
        sheet.cell(row=cnt, column=7).border = thin_border
        sheet.cell(row=cnt, column=8).border = thin_border
        sheet.cell(row=cnt, column=9).border = thin_border
        sheet.cell(row=cnt, column=10).border = thin_border
        sheet.cell(row=cnt, column=11).border = thin_border
        cnt += 1

    for i in range(5, order.count()+5):
        sheet.cell(row=i, column=2).border = thin_border
        sheet.cell(row=i, column=3).border = thin_border
        sheet.cell(row=i, column=4).border = thin_border
        sheet.cell(row=i, column=5).border = thin_border

    sheet["N4"].value = str(company)
    sheet["N5"].value = f'{date_before} - {date_after}'
    sheet["N6"].value = int(len(unique_unit_order))
    sheet["N7"].value = int(all_cnt_component)
    sheet["N8"].value = int(all_summ_invoice)
    sheet["N10"].value = float(min_invoice)
    sheet["N11"].value = float(max_invoice)
    sheet["N12"].value = float(avg_invoice)

    sheet["N4"].alignment = Alignment(horizontal='center')
    sheet["N5"].alignment = Alignment(horizontal='center')
    sheet["N6"].alignment = Alignment(horizontal='center')
    sheet["N7"].alignment = Alignment(horizontal='center')
    sheet["N8"].alignment = Alignment(horizontal='center')
    sheet["N10"].alignment = Alignment(horizontal='center')
    sheet["N11"].alignment = Alignment(horizontal='center')
    sheet["N12"].alignment = Alignment(horizontal='center')

    path_save = os.path.join(
        os.getcwd(),
        "base/Order_report/report.xlsx"
    )
    wb.save(path_save)
    
